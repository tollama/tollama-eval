"""Data export utilities for benchmark results."""

from __future__ import annotations

from pathlib import Path

from ts_autopilot.contracts import BenchmarkResult
from ts_autopilot.logging_config import get_logger

logger = get_logger("export")


def _build_per_series_winner_summary(
    result: BenchmarkResult,
) -> tuple[list[str], list[dict]]:
    """Build average per-series winner rows across all models."""
    scores_by_model: dict[str, dict[str, list[float]]] = {}
    ordered_models = [entry.name for entry in result.leaderboard]

    for model in result.models:
        if model.name not in ordered_models:
            ordered_models.append(model.name)
        model_scores: dict[str, list[float]] = {}
        for fold in model.folds:
            for sid, score in fold.series_scores.items():
                model_scores.setdefault(sid, []).append(score)
        if model_scores:
            scores_by_model[model.name] = model_scores

    comparable_rows: list[dict] = []
    all_series = sorted(
        {
            sid
            for model_scores in scores_by_model.values()
            for sid in model_scores
        }
    )

    for sid in all_series:
        averages: dict[str, float] = {}
        for model_name in ordered_models:
            series_scores = scores_by_model.get(model_name, {}).get(sid)
            if series_scores:
                averages[model_name] = round(
                    sum(series_scores) / len(series_scores),
                    4,
                )

        if len(averages) < 2:
            continue

        ranked = sorted(averages.items(), key=lambda item: item[1])
        winner_name, winner_score = ranked[0]
        runner_up_name, runner_up_score = ranked[1]
        comparable_rows.append(
            {
                "series": sid,
                "winner": winner_name,
                "winner_mase": winner_score,
                "runner_up": runner_up_name,
                "runner_up_mase": runner_up_score,
                "margin": round(runner_up_score - winner_score, 4),
                "spread": round(ranked[-1][1] - winner_score, 4),
                "scores": averages,
            }
        )

    comparable_rows.sort(
        key=lambda row: (-row["winner_mase"], row["margin"], row["series"])
    )
    return ordered_models, comparable_rows


def export_leaderboard_csv(result: BenchmarkResult, output_path: str | Path) -> Path:
    """Export leaderboard as a CSV file.

    Args:
        result: Benchmark result with leaderboard data.
        output_path: Path for the output CSV file.

    Returns:
        Path to the written CSV file.
    """
    import pandas as pd

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = []
    for entry in result.leaderboard:
        rows.append(
            {
                "rank": entry.rank,
                "model": entry.name,
                "mase": entry.mean_mase,
                "smape": entry.mean_smape,
                "rmsse": entry.mean_rmsse,
                "mae": entry.mean_mae,
            }
        )

    df = pd.DataFrame(rows)
    df.to_csv(output_path, index=False)
    logger.info("Exported leaderboard CSV: %s", output_path)
    return output_path


def export_fold_details_csv(result: BenchmarkResult, output_path: str | Path) -> Path:
    """Export per-model per-fold details as a CSV file.

    Args:
        result: Benchmark result with model data.
        output_path: Path for the output CSV file.

    Returns:
        Path to the written CSV file.
    """
    import pandas as pd

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = []
    for model in result.models:
        for fold in model.folds:
            rows.append(
                {
                    "model": model.name,
                    "fold": fold.fold,
                    "cutoff": fold.cutoff,
                    "mase": fold.mase,
                    "smape": fold.smape,
                    "rmsse": fold.rmsse,
                    "mae": fold.mae,
                }
            )

    df = pd.DataFrame(rows)
    df.to_csv(output_path, index=False)
    logger.info("Exported fold details CSV: %s", output_path)
    return output_path


def export_per_series_csv(result: BenchmarkResult, output_path: str | Path) -> Path:
    """Export per-series scores across all models as a CSV file.

    Args:
        result: Benchmark result with per-series score data.
        output_path: Path for the output CSV file.

    Returns:
        Path to the written CSV file.
    """
    import pandas as pd

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = []
    for model in result.models:
        for fold in model.folds:
            for sid, score in fold.series_scores.items():
                rows.append(
                    {
                        "model": model.name,
                        "fold": fold.fold,
                        "series": sid,
                        "mase": score,
                    }
                )

    df = pd.DataFrame(rows)
    df.to_csv(output_path, index=False)
    logger.info("Exported per-series CSV: %s", output_path)
    return output_path


def export_per_series_winners_csv(
    result: BenchmarkResult,
    output_path: str | Path,
) -> Path:
    """Export average per-series winners and competition margins as a CSV file."""
    import pandas as pd

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    ordered_models, rows = _build_per_series_winner_summary(result)

    flattened_rows = []
    for row in rows:
        flat_row = {
            "series": row["series"],
            "winner": row["winner"],
            "winner_mase": row["winner_mase"],
            "runner_up": row["runner_up"],
            "runner_up_mase": row["runner_up_mase"],
            "margin": row["margin"],
            "spread": row["spread"],
        }
        for model_name in ordered_models:
            flat_row[f"{model_name}_mase"] = row["scores"].get(model_name)
        flattened_rows.append(flat_row)

    df = pd.DataFrame(flattened_rows)
    df.to_csv(output_path, index=False)
    logger.info("Exported per-series winners CSV: %s", output_path)
    return output_path


def export_excel(result: BenchmarkResult, output_path: str | Path) -> Path:
    """Export benchmark results as a formatted Excel workbook.

    Creates sheets: Executive Summary, Leaderboard, Fold Details,
    Per-Series Scores, Per-Series Winners, Data Profile.

    Requires openpyxl. Install with: pip install "tollama-eval[excel]"

    Args:
        result: Benchmark result.
        output_path: Path for the output .xlsx file.

    Returns:
        Path to the written Excel file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.info(
            'openpyxl not installed. Install with: pip install "tollama-eval[excel]"'
        )
        raise

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # --- Styles ---
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1E40AF", end_color="1E40AF", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    good_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    bad_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
    good_font = Font(color="059669", bold=True)
    bad_font = Font(color="DC2626", bold=True)
    title_font = Font(bold=True, size=14, color="1E40AF")

    def _style_header(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        row: int,
        cols: int,
    ) -> None:
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _auto_width(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # --- Sheet 1: Executive Summary ---
    ws_summary = wb.active
    ws_summary.title = "Executive Summary"
    from ts_autopilot.reporting.executive_summary import generate_executive_summary

    summary = generate_executive_summary(result)
    row = 1
    ws_summary.cell(row=row, column=1, value="Benchmark Report").font = title_font
    row += 2
    ws_summary.cell(row=row, column=1, value="Overview").font = Font(bold=True, size=12)
    row += 1
    ws_summary.cell(row=row, column=1, value=summary.overview)
    ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    if summary.winner:
        cell = ws_summary.cell(row=row, column=1, value="Winner")
        cell.font = Font(bold=True, size=12)
        row += 1
        ws_summary.cell(row=row, column=1, value=summary.winner)
        ws_summary.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2

    if summary.key_findings:
        cell = ws_summary.cell(row=row, column=1, value="Key Findings")
        cell.font = Font(bold=True, size=12)
        row += 1
        for finding in summary.key_findings:
            ws_summary.cell(row=row, column=1, value=f"  - {finding}")
            row += 1
        row += 1

    if summary.recommendations:
        cell = ws_summary.cell(row=row, column=1, value="Recommendations")
        cell.font = Font(bold=True, size=12)
        row += 1
        for i, rec in enumerate(summary.recommendations, 1):
            ws_summary.cell(row=row, column=1, value=f"  {i}. {rec}")
            row += 1

    ws_summary.column_dimensions["A"].width = 100

    # --- Sheet 2: Leaderboard ---
    ws_lb = wb.create_sheet("Leaderboard")
    headers = ["Rank", "Model", "MASE", "SMAPE (%)", "RMSSE", "MAE"]
    for col, h in enumerate(headers, 1):
        ws_lb.cell(row=1, column=col, value=h)
    _style_header(ws_lb, 1, len(headers))

    for i, entry in enumerate(result.leaderboard, 2):
        rank_cell = ws_lb.cell(row=i, column=1, value=entry.rank)
        rank_cell.alignment = Alignment(horizontal="center")
        ws_lb.cell(row=i, column=2, value=entry.name)
        mase_cell = ws_lb.cell(row=i, column=3, value=round(entry.mean_mase, 4))
        ws_lb.cell(row=i, column=4, value=round(entry.mean_smape, 2))
        rmsse_cell = ws_lb.cell(row=i, column=5, value=round(entry.mean_rmsse, 4))
        ws_lb.cell(row=i, column=6, value=round(entry.mean_mae, 4))
        for col in range(1, 7):
            ws_lb.cell(row=i, column=col).border = thin_border
        # Conditional formatting
        if entry.mean_mase < 1.0:
            mase_cell.fill = good_fill
            mase_cell.font = good_font
        elif entry.mean_mase > 1.0:
            mase_cell.fill = bad_fill
            mase_cell.font = bad_font
        if entry.mean_rmsse < 1.0:
            rmsse_cell.fill = good_fill
            rmsse_cell.font = good_font
        elif entry.mean_rmsse > 1.0:
            rmsse_cell.fill = bad_fill
            rmsse_cell.font = bad_font

    _auto_width(ws_lb)

    # --- Sheet 3: Fold Details ---
    ws_folds = wb.create_sheet("Fold Details")
    fold_headers = ["Model", "Fold", "Cutoff", "MASE", "SMAPE (%)", "RMSSE", "MAE"]
    for col, h in enumerate(fold_headers, 1):
        ws_folds.cell(row=1, column=col, value=h)
    _style_header(ws_folds, 1, len(fold_headers))

    row = 2
    for model in result.models:
        for fold in model.folds:
            ws_folds.cell(row=row, column=1, value=model.name)
            ws_folds.cell(row=row, column=2, value=fold.fold)
            ws_folds.cell(row=row, column=3, value=fold.cutoff)
            ws_folds.cell(row=row, column=4, value=round(fold.mase, 4))
            ws_folds.cell(row=row, column=5, value=round(fold.smape, 2))
            ws_folds.cell(row=row, column=6, value=round(fold.rmsse, 4))
            ws_folds.cell(row=row, column=7, value=round(fold.mae, 4))
            for col in range(1, 8):
                ws_folds.cell(row=row, column=col).border = thin_border
            row += 1

    _auto_width(ws_folds)

    # --- Sheet 4: Per-Series Scores ---
    ws_series = wb.create_sheet("Per-Series Scores")
    series_headers = ["Model", "Fold", "Series", "MASE"]
    for col, h in enumerate(series_headers, 1):
        ws_series.cell(row=1, column=col, value=h)
    _style_header(ws_series, 1, len(series_headers))

    row = 2
    for model in result.models:
        for fold in model.folds:
            for sid, score in fold.series_scores.items():
                ws_series.cell(row=row, column=1, value=model.name)
                ws_series.cell(row=row, column=2, value=fold.fold)
                ws_series.cell(row=row, column=3, value=sid)
                ws_series.cell(row=row, column=4, value=round(score, 6))
                for col in range(1, 5):
                    ws_series.cell(row=row, column=col).border = thin_border
                row += 1

    _auto_width(ws_series)

    # --- Sheet 5: Per-Series Winners ---
    ws_winners = wb.create_sheet("Per-Series Winners")
    ordered_models, winner_rows = _build_per_series_winner_summary(result)
    winner_headers = [
        "Series",
        "Winner",
        "Winner MASE",
        "Runner-up",
        "Runner-up MASE",
        "Margin",
        "Spread",
        *[f"{model_name} MASE" for model_name in ordered_models],
    ]
    for col, header in enumerate(winner_headers, 1):
        ws_winners.cell(row=1, column=col, value=header)
    _style_header(ws_winners, 1, len(winner_headers))

    row = 2
    for winner_row in winner_rows:
        ws_winners.cell(row=row, column=1, value=winner_row["series"])
        ws_winners.cell(row=row, column=2, value=winner_row["winner"])
        winner_mase_cell = ws_winners.cell(
            row=row,
            column=3,
            value=winner_row["winner_mase"],
        )
        ws_winners.cell(row=row, column=4, value=winner_row["runner_up"])
        ws_winners.cell(row=row, column=5, value=winner_row["runner_up_mase"])
        ws_winners.cell(row=row, column=6, value=winner_row["margin"])
        ws_winners.cell(row=row, column=7, value=winner_row["spread"])

        for idx, model_name in enumerate(ordered_models, start=8):
            score = winner_row["scores"].get(model_name)
            cell = ws_winners.cell(row=row, column=idx, value=score)
            if score is not None:
                if score < 1.0:
                    cell.fill = good_fill
                    cell.font = good_font
                elif score > 1.0:
                    cell.fill = bad_fill
                    cell.font = bad_font

        if winner_row["winner_mase"] < 1.0:
            winner_mase_cell.fill = good_fill
            winner_mase_cell.font = good_font
        elif winner_row["winner_mase"] > 1.0:
            winner_mase_cell.fill = bad_fill
            winner_mase_cell.font = bad_font

        for col in range(1, len(winner_headers) + 1):
            ws_winners.cell(row=row, column=col).border = thin_border
        row += 1

    _auto_width(ws_winners)

    # --- Sheet 6: Data Profile ---
    ws_profile = wb.create_sheet("Data Profile")
    profile_data = [
        ("Series Count", result.profile.n_series),
        ("Total Rows", result.profile.total_rows),
        ("Frequency", result.profile.frequency),
        ("Season Length", result.profile.season_length_guess),
        ("Min Series Length", result.profile.min_length),
        ("Max Series Length", result.profile.max_length),
        ("Missing Ratio", f"{result.profile.missing_ratio:.2%}"),
        ("Forecast Horizon", result.config.horizon),
        ("CV Folds", result.config.n_folds),
    ]
    ws_profile.cell(row=1, column=1, value="Property").font = Font(bold=True, size=11)
    ws_profile.cell(row=1, column=2, value="Value").font = Font(bold=True, size=11)
    _style_header(ws_profile, 1, 2)

    for i, (prop, val) in enumerate(profile_data, 2):
        ws_profile.cell(row=i, column=1, value=prop)
        ws_profile.cell(row=i, column=2, value=val)
        ws_profile.cell(row=i, column=1).border = thin_border
        ws_profile.cell(row=i, column=2).border = thin_border

    _auto_width(ws_profile)

    # Save
    wb.save(str(output_path))
    logger.info("Exported Excel workbook: %s", output_path)
    return output_path
